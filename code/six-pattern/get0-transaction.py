import csv
import os
from openpyxl import load_workbook, Workbook
from collections import defaultdict
import time
def match_address_and_export(csv_file_path, xlsx_file_path, output_folder):
    start_time = time.time()
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    wb = load_workbook(xlsx_file_path, read_only=True)
    ws = wb.active
    # ws = wb['地址重合-前-合并']
    addresses = {row[0] for row in ws.iter_rows(min_row=1, max_col=1, values_only=True)}
    wb.close()
    matched_addresses = defaultdict(list)
    # 遍历
    with open(csv_file_path, mode='r', encoding='utf-8', newline='') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            from_address = row['from']
            to_address = row['to']
            if from_address in addresses:
                matched_addresses[from_address].append((row, 0))
            if to_address in addresses:
                matched_addresses[to_address].append((row, 1))
    # 对于每个匹配成功的地址，检查是否存在同名xlsx文件来决定是否新建或续写
    for address, rows in matched_addresses.items():
        sanitized_address = ''.join(c for c in address if c.isalnum())
        file_path = os.path.join(output_folder, f'{sanitized_address}.xlsx')
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            worksheet = workbook.active
        else:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(list(reader.fieldnames) + ['标记'])
        for row, mark in rows:
            worksheet.append(list(row.values()) + [mark])
        workbook.save(file_path)
        workbook.close()
    end_time = time.time()
    print(xlsx_file,'处理完成', csv_file, '耗时:', end_time - start_time, '秒')
# 账本CSV文件列表
csv_files = [
    'data/9000000to9999999_BlockTransaction.csv',
    'data/10000000to10999999_BlockTransaction.csv',
    'data/11000000to11999999_BlockTransaction.csv',
    'data/12000000to12999999_BlockTransaction.csv',
    'data/13000000to13249999_BlockTransaction.csv',
    'data/13250000to13499999_BlockTransaction.csv',
    'data/13500000to13749999_BlockTransaction.csv',
    'data/13750000to13999999_BlockTransaction.csv',
    'data/14000000to14249999_BlockTransaction.csv',
    'data/14500000to14749999_BlockTransaction.csv',
    'data/14750000to14999999_BlockTransaction.csv',
    'data/15000000to15249999_BlockTransaction.csv',
    'data/15250000to15499999_BlockTransaction.csv',
]
xlsx_files = [
    'dataset/TC合约.xlsx',
]

output_folder = 'dataset/0阶交易'

# 遍历CSV文件，并且对每个文件调用匹配函数
# for csv_file in csv_files:
#     match_address_and_export(csv_file, xlsx_file, output_folder)

for xlsx_file in xlsx_files:
    for csv_file in csv_files:
        match_address_and_export(csv_file, xlsx_file, output_folder)